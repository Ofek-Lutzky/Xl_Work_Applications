from openpyxl import Workbook, load_workbook
from tkinter import *
from datetime import date


Path = 'python_work_places.xlsx'


#just for the first time
# def InitializeExcel(path):
#     excelWBook = load_workbook(path)
#     # if excelWBook != None:
#     #     return
#
#     excelWSheet = excelWBook.active
#     print(excelWSheet.values)
#     # excelWSheet.title = "python_work_places"
#     # print(excelWSheet['B1'].value) # just a Test
#     excelWSheet.append(['Date', "Company name", "Position", "City", "Answered" , "intreview"])
#     excelWBook.save(path)
# #
# InitializeExcel('python_work_places.xlsx') # init the headlines

def sheetData(path):
    excelWBook = load_workbook(path)
    excelWSheet = excelWBook.active
    Data = []
    counter = 1
    for row in excelWSheet.iter_rows(values_only=True):
        # print(row)
        line = ""
        if counter == 1:
            for i in row:
                if i == None:
                    line = line + "None" + "      "
                else:
                    line = line + i + "      "
            counter+=1
        else:
            for i in row:
                if i == None:
                    line = line + "None" + "           "
                else:
                    line = line + str(i) + "           "
        Data.append(line)


    return Data



def AddWorkPlace(path,company,Position,City,Ans = "No",interview= "No"):


    if (company != '' and Position != '' and City != ''):
        label.config(text = "Added " + company +" " + Position +" " + City)
        excelWBook = load_workbook(path)
        excelWSheet = excelWBook.active
        excelWSheet.append([date.today().strftime("%d/%m/%Y"),company,Position,City,Ans,interview])
        excelWBook.save(path)

    else:
        label.config(text="Didn't enter: Company,Position,City!")


#funtion to remove the last index

# def DeleteLastWork(path):
#     excelWBook = load_workbook(path)
#     excelWSheet = excelWBook.active
#
#     if excelWSheet.max_row == 1: # that it woun't delete the headers of the table
#         return
#
#     excelWSheet.delete_rows(idx = excelWSheet.max_row)


# helper func to find the index of the row to update or delete
def GetIndex(path,company, position, city):
    excelWBook = load_workbook(path)
    excelWSheet = excelWBook.active
    counter = 1
    for row in excelWSheet.iter_rows(values_only=True):
        if (company == row[1] and position == row[2] and city == row[3]):
            return counter
        counter+=1
    return -1


def DeleteTheChosen(path,company, position, city):
    excelWBook = load_workbook(path)
    excelWSheet = excelWBook.active
    counter = 1
    flag = 0

    if (company == "Company name" and position == "Position" and city == "City"):
        return

    flag = GetIndex(path,company, position, city)
    if flag == -1:
        return
    excelWSheet.delete_rows(idx=flag,amount = 1)
    excelWBook.save(path)


def Update(path,company, position, city,answer,inerview):
    excelWBook = load_workbook(path)
    excelWSheet = excelWBook.active

    flag = GetIndex(path, company, position, city)
    if flag == -1:
        return

    if answer != None:
        excelWSheet.cell(row=flag,column=5).value = answer
    if inerview != None:
        excelWSheet.cell(row=flag, column=6).value = inerview

    excelWBook.save(path)


#gui function

window = Tk()

#headlines
l1 = Label(window, text = "Company Name:")
l1.grid(row = 0 , column = 0)
l1 = Label(window, text="Position:")
l1.grid(row=1, column=0)
l1 = Label(window, text="City:")
l1.grid(row=2, column=0)
l1 = Label(window, text="Ans:")
l1.grid(row=0, column=2)
l1 = Label(window, text="Interview:")
l1.grid(row=1, column=2)


#getting the input
Company_name = StringVar()
e1 = Entry(window, textvariable = Company_name)
e1.grid(row = 0,column = 1)

Position_name = StringVar()
e2 = Entry(window, textvariable=Position_name)
e2.grid(row=1, column=1)

City_name = StringVar()
e3 = Entry(window, textvariable=City_name)
e3.grid(row=2, column=1)

Ans_text = StringVar()
e4 = Entry(window, textvariable=Ans_text)
e4.grid(row=0, column=3)

Interview_time = StringVar()
e5 = Entry(window, textvariable=Interview_time)
e5.grid(row=1, column=3)

# define listBox
listBox = Listbox(window, height = 10, width = 100)
listBox.grid(row = 3, column = 0, columnspan = 2, rowspan = 5)

#Attach scroll bar to list
sb1 = Scrollbar(window)
sb1.grid(row = 4, column = 2, rowspan = 5)

listBox.configure(yscrollcommand = sb1.set)
sb1.configure(command = listBox.yview)



# all the buttons functions

def b1F():
    Data = sheetData(Path)
    for i in Data:
        listBox.insert(END, i + "\n")


#label init
global label
label = Label(window, text='')
label.grid(row = 9 , column = 0)


def b2F():
    AddWorkPlace(Path, e1.get(), e2.get(), e3.get(), e4.get(), e5.get())



def b3F():
    lineToUpdate = listBox.get(ANCHOR).split()
    company = lineToUpdate[1]
    position = lineToUpdate[2]
    city = lineToUpdate[3]

    DeleteTheChosen(Path,company,position,city)

def b4F():
    lineToUpdate = listBox.get(ANCHOR).split()
    company = lineToUpdate[1]
    position = lineToUpdate[2]
    city = lineToUpdate[3]


    Update(Path,company, position, city, e4.get(),e5.get())



def b5F():
    listBox.delete(0, END)

#Define buttons
b1 = Button(window,text="View All", width = 15, command= b1F)
b1.grid(row=3, column=3)

b2 = Button(window, text="Add Position", width=15, command= b2F)
b2.grid(row=4, column=3)

b3 = Button(window, text="Delete Chosen", width=15, command= b3F)
b3.grid(row=5, column=3)

b4 = Button(window, text="Update Position", width=15, command= b4F)
b4.grid(row=6, column=3)

b5 = Button(window, text="Delete Board", width=15, command= b5F)
b5.grid(row=7, column=3)




window.mainloop()



